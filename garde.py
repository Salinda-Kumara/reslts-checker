import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import pandas as pd
import openpyxl
import os
import webbrowser
import tempfile
import re
import math
from datetime import datetime

# Set Appearance and Theme
ctk.set_appearance_mode("Dark")  # Modes: "System" (standard), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

class StudentResultApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Window Setup
        self.title("Student Result System - SAB Campus")
        self.geometry("1300x850")
        
        # Variables
        self.all_sheets_data = {} 
        self.file_path = None
        self.current_sheet_name = None
        self.current_student_idx = None
        self.subject_columns_per_sheet = {}
        self.subject_credits_per_sheet = {}  # Store credit values for each subject
        self.include_gpa_class = True  # Toggle for including GPA and Class in transcript
        self.current_subject_rows = []
        self.subject_results_window = None
        
        # Trackers
        self.pending_changes = {} 
        self.pending_deletes = {} 

        # Layout Configuration
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # --- SIDEBAR ---
        self.sidebar_frame = ctk.CTkFrame(self, width=200, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(7, weight=1)



        self.logo_label = ctk.CTkLabel(self.sidebar_frame, text="Result System", font=ctk.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.btn_load = ctk.CTkButton(self.sidebar_frame, text="Load Workbook", command=self.load_workbook, fg_color="#2E7D32", hover_color="#1B5E20")
        self.btn_load.grid(row=1, column=0, padx=20, pady=10)

        self.btn_save = ctk.CTkButton(self.sidebar_frame, text="Save Changes", command=self.smart_save, state="disabled", fg_color="#C62828", hover_color="#B71C1C")
        self.btn_save.grid(row=2, column=0, padx=20, pady=10)

        self.btn_export_sheet = ctk.CTkButton(self.sidebar_frame, text="Export Sheet", command=self.print_full_sheet, state="disabled", fg_color="#455A64", hover_color="#37474F")
        self.btn_export_sheet.grid(row=3, column=0, padx=20, pady=10)
        
        # Toggle for GPA and Class in Transcript
        self.toggle_frame = ctk.CTkFrame(self.sidebar_frame, fg_color="transparent")
        self.toggle_frame.grid(row=4, column=0, padx=20, pady=10, sticky="ew")
        
        self.lbl_toggle = ctk.CTkLabel(self.toggle_frame, text="Include GPA & Class", font=ctk.CTkFont(size=12))
        self.lbl_toggle.pack(side="left", padx=(0, 10))
        
        self.toggle_gpa_class = ctk.CTkSwitch(self.toggle_frame, text="", command=self.toggle_gpa_class_callback)
        self.toggle_gpa_class.pack(side="left")
        self.toggle_gpa_class.select()  # Default ON
        
        # GPA Display in Sidebar
        self.gpa_frame = ctk.CTkFrame(self.sidebar_frame, fg_color="transparent")
        self.gpa_frame.grid(row=5, column=0, padx=20, pady=(20, 10), sticky="ew")
        
        self.lbl_gpa_title = ctk.CTkLabel(self.gpa_frame, text="GPA", font=ctk.CTkFont(size=16, weight="bold"))
        self.lbl_gpa_title.pack()
        
        self.lbl_gpa_value = ctk.CTkLabel(self.gpa_frame, text="-", font=ctk.CTkFont(size=32, weight="bold"), text_color="#4CAF50")
        self.lbl_gpa_value.pack()

        # Class Awarded Display
        self.class_frame = ctk.CTkFrame(self.sidebar_frame, fg_color="transparent")
        self.class_frame.grid(row=6, column=0, padx=20, pady=(5, 10), sticky="ew")
        
        self.lbl_class_title = ctk.CTkLabel(self.class_frame, text="Class Awarded", font=ctk.CTkFont(size=16, weight="bold"))
        self.lbl_class_title.pack()
        
        self.lbl_class_value = ctk.CTkLabel(self.class_frame, text="-", font=ctk.CTkFont(size=18), text_color="#FFC107", wraplength=180)
        self.lbl_class_value.pack()

        # Appearance Mode Switch
        self.appearance_mode_label = ctk.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=8, column=0, padx=20, pady=(10, 0))

        self.appearance_mode_optionemenu = ctk.CTkOptionMenu(self.sidebar_frame, values=["Dark", "Light", "System"], command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=9, column=0, padx=20, pady=(10, 20))


        # --- MAIN CONTENT ---
        self.main_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.main_frame.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        
        # Search Area
        self.search_frame = ctk.CTkFrame(self.main_frame)
        self.search_frame.pack(fill="x", pady=(0, 15))
        
        self.search_var = ctk.StringVar(value="")
        self.entry_search = ctk.CTkEntry(
            self.search_frame,
            placeholder_text="Search Registration Number...",
            placeholder_text_color="#B0BEC5",
            width=300,
            textvariable=self.search_var
        )
        self.entry_search.pack(side="left", padx=15, pady=15)
        self.entry_search.bind('<Return>', lambda event: self.search_global())
        
        self.btn_search = ctk.CTkButton(self.search_frame, text="Search", width=100, command=self.search_global)
        self.btn_search.pack(side="left", padx=5, pady=15)

        self.subject_search_var = ctk.StringVar(value="")
        self.entry_subject_search = ctk.CTkEntry(
            self.search_frame,
            placeholder_text="Search Subject...",
            placeholder_text_color="#B0BEC5",
            width=250,
            textvariable=self.subject_search_var,
            state="disabled"
        )
        self.entry_subject_search.pack(side="left", padx=5, pady=15)
        self.entry_subject_search.bind('<Return>', lambda event: self.filter_subjects())
        self.entry_subject_search.bind('<KeyRelease>', lambda event: self.filter_subjects())

        self.btn_subject_filter = ctk.CTkButton(self.search_frame, text="Filter Subject", width=90, command=self.filter_subjects, state="disabled")
        self.btn_subject_filter.pack(side="left", padx=(0, 10), pady=15)
        
        self.lbl_status = ctk.CTkLabel(self.search_frame, text="", text_color="gray")
        self.lbl_status.pack(side="left", padx=15, pady=15)
        
        self.lbl_current_sheet = ctk.CTkLabel(self.search_frame, text="Sheet: -", font=ctk.CTkFont(weight="bold"))
        self.lbl_current_sheet.pack(side="right", padx=15, pady=15)

        # Info Card
        self.info_frame = ctk.CTkFrame(self.main_frame)
        self.info_frame.pack(fill="x", pady=(0, 15))
        
        self.lbl_name = ctk.CTkLabel(self.info_frame, text="Name: -", font=ctk.CTkFont(size=16, weight="bold"))
        self.lbl_name.pack(anchor="w", padx=15, pady=(10, 5))
        
        self.lbl_reg = ctk.CTkLabel(self.info_frame, text="Reg No: -", font=ctk.CTkFont(size=14))
        self.lbl_reg.pack(anchor="w", padx=15, pady=(0, 10))

        # Results Table (Treeview)
        self.table_frame = ctk.CTkFrame(self.main_frame)
        self.table_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # Style Treeview
        self.style = ttk.Style()
        self.style.theme_use("default")
        self.apply_treeview_style(ctk.get_appearance_mode())

        self.tree_scroll = ctk.CTkScrollbar(self.table_frame)
        self.tree_scroll.pack(side="right", fill="y")
        
        columns = ("SNo", "Subject", "Grade", "Points")
        self.tree = ttk.Treeview(self.table_frame, columns=columns, show="headings", yscrollcommand=self.tree_scroll.set)
        self.tree.heading("SNo", text="#")
        self.tree.heading("Subject", text="Subject Name")
        self.tree.heading("Grade", text="Grade")
        self.tree.heading("Points", text="Points")
        
        self.tree.column("SNo", width=60, anchor="center")
        self.tree.column("Subject", width=500)
        self.tree.column("Grade", width=100, anchor="center")
        self.tree.column("Points", width=100, anchor="center")
        
        self.tree.pack(fill="both", expand=True, padx=2, pady=2)
        self.tree_scroll.configure(command=self.tree.yview)

        # Action Buttons
        self.action_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.action_frame.pack(fill="x")
        
        self.btn_edit = ctk.CTkButton(self.action_frame, text="Edit Grade", command=self.edit_grade, state="disabled", fg_color="#FFA000", hover_color="#FF6F00")
        self.btn_edit.pack(side="left", padx=(0, 10))
        
        self.btn_print = ctk.CTkButton(self.action_frame, text="Print Transcript", command=self.print_student_transcript, state="disabled", fg_color="#1976D2", hover_color="#0D47A1")
        self.btn_print.pack(side="left", padx=10)
        
        self.btn_subject_results = ctk.CTkButton(
            self.action_frame,
            text="View Subject Results",
            command=self.open_subject_results_window,
            state="disabled",
            fg_color="#00897B",
            hover_color="#00695C"
        )
        self.btn_subject_results.pack(side="left", padx=10)
        
        self.btn_delete = ctk.CTkButton(self.action_frame, text="Delete Student", command=self.delete_student, state="disabled", fg_color="#D32F2F", hover_color="#B71C1C")
        self.btn_delete.pack(side="right")

    def toggle_gpa_class_callback(self):
        """Toggle callback for including GPA and Class in transcript"""
        self.include_gpa_class = self.toggle_gpa_class.get()
    
    def change_appearance_mode_event(self, new_appearance_mode: str):
        ctk.set_appearance_mode(new_appearance_mode)
        self.apply_treeview_style(new_appearance_mode)

    def apply_treeview_style(self, mode):
        if mode == "Light":
            bg_color = "#ffffff"
            fg_color = "#000000"
            field_bg = "#ffffff"
            heading_bg = "#e1e1e1"
            heading_fg = "#000000"
            selected_bg = "#3a7ebf"
        else:
            bg_color = "#2b2b2b"
            fg_color = "white"
            field_bg = "#2b2b2b"
            heading_bg = "#1f1f1f"
            heading_fg = "white"
            selected_bg = "#1f6aa5"

        self.style.configure("Treeview", 
                        background=bg_color, 
                        foreground=fg_color, 
                        fieldbackground=field_bg, 
                        rowheight=30,
                        font=("Arial", 11))
        self.style.configure("Treeview.Heading", 
                        background=heading_bg, 
                        foreground=heading_fg, 
                        font=("Arial", 11, "bold"))
        self.style.map("Treeview", background=[('selected', selected_bg)])

    def load_workbook(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if not file_path: return
            
        try:
            all_sheets_raw = pd.read_excel(file_path, sheet_name=None, header=None)
            self.all_sheets_data = {}
            self.subject_columns_per_sheet = {}
            all_reg_numbers = []
            
            for sheet_name, raw_df in all_sheets_raw.items():
                if len(raw_df) < 1: continue
                
                # Find Header Row
                header_row_idx = -1
                for i in range(min(10, len(raw_df))):
                    row_vals = raw_df.iloc[i].astype(str).str.lower().tolist()
                    if any("registration" in x for x in row_vals) and (any("no" in x for x in row_vals) or any("number" in x for x in row_vals)):
                        header_row_idx = i
                        break
                
                if header_row_idx == -1: continue
                    
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx)
                
                # Merge Subject Row (Row below header)
                if len(df) > 0:
                    subject_row = df.iloc[0]
                    new_columns = list(df.columns)
                    for i in range(len(new_columns)):
                        if i >= 4: 
                            val = subject_row.iloc[i]
                            if pd.notna(val) and str(val).strip() != "":
                                new_columns[i] = str(val).strip()
                    df.columns = new_columns
                    df = df.drop(0).reset_index(drop=True)

                # Normalize Columns
                df.columns = [str(c).strip() for c in df.columns]
                reg_col_found = False
                for col in df.columns:
                    if "registration" in col.lower() or "reg no" in col.lower():
                        df.rename(columns={col: 'Registration Number'}, inplace=True)
                        reg_col_found = True
                        break
                
                if not reg_col_found: continue
                
                # --- FILTERING LOGIC ---
                exclude_keywords = [
                    'serial', 'registration', 'name', 'general', 'special', 'drop', 'batch', 'transfer',
                    'finance clearance', 'results confirmation', 'gpa', 'class', 'effective', 'total no', 
                    'nq ese', 'ab ese', 'repeat ese', 'hold', 'results released', 'range of marks', 'grade points', 'letter grade',
                    'pendings', 'pending'
                ]
                grade_cols = ['A+', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'E', 'AB', 'B and Above', 'C and above', 'A+/A/A-', 'EX', '-', 'C-/D+']
                
                subject_cols = []
                for col in df.columns:
                    cl = col.lower()
                    if cl == 's': continue
                    if any(k in cl for k in exclude_keywords): continue
                    if col in grade_cols: continue
                    if "unnamed" in cl: continue
                    subject_cols.append(col)
                
                self.subject_columns_per_sheet[sheet_name] = subject_cols
                
                # Extract credit values from subject code (last digit of subject code)
                # Example: "BSAA 11013" -> credit = 3
                subject_credits = {}
                for col_name in subject_cols:
                    # Extract credit from subject code - last digit
                    subject_code = str(col_name).strip()
                    # Find all digits in the subject code
                    digits = re.findall(r'\d', subject_code)
                    if digits:
                        # Get the last digit as credit value
                        credit_val = int(digits[-1])
                        if credit_val > 0:  # Valid credit should be > 0
                            subject_credits[col_name] = float(credit_val)
                
                self.subject_credits_per_sheet[sheet_name] = subject_credits
                self.all_sheets_data[sheet_name] = df
                all_reg_numbers.extend(df['Registration Number'].dropna().astype(str).tolist())

            self.file_path = file_path
            unique_regs = sorted(list(set(all_reg_numbers)))
            
            self.lbl_status.configure(text=f"Loaded {len(unique_regs)} students.")
            self.pending_changes = {}
            self.pending_deletes = {}
            self.btn_save.configure(state="normal")
            self.btn_export_sheet.configure(state="normal")
            self.btn_subject_results.configure(state="normal")
            self.clear_ui()
            
            messagebox.showinfo("Success", f"Loaded {len(self.all_sheets_data)} valid sheets.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not load workbook: {e}")

    def search_global(self):
        if not self.all_sheets_data: return
        search_term = self.search_var.get().strip()
        if not search_term: return

        all_matches = []
        for sheet_name, df in self.all_sheets_data.items():
            mask = df['Registration Number'].astype(str).str.contains(search_term, case=False, na=False)
            matches = df[mask]
            for idx, row in matches.iterrows():
                all_matches.append((sheet_name, idx, row))

        if not all_matches:
            messagebox.showerror("Not Found", "Student not found.")
            self.clear_ui()
        elif len(all_matches) == 1:
            self.load_student_into_ui(*all_matches[0][:2])
        else:
            self.show_selection_popup(all_matches)

    def show_selection_popup(self, matches):
        popup = ctk.CTkToplevel(self)
        popup.title("Select Student")
        popup.geometry("600x400")
        
        # Make it modal
        popup.transient(self)
        popup.grab_set()
        
        lbl = ctk.CTkLabel(popup, text="Multiple matches found. Select one:", font=ctk.CTkFont(size=14, weight="bold"))
        lbl.pack(pady=10)

        scroll_frame = ctk.CTkScrollableFrame(popup)
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)

        for sheet, idx, row in matches:
            name_col = next((c for c in row.index if "name" in c.lower()), "Name")
            name_val = row[name_col] if name_col in row else "Unknown"
            reg_val = row['Registration Number']
            
            btn_text = f"[{sheet}] {reg_val} | {name_val}"
            btn = ctk.CTkButton(scroll_frame, text=btn_text, anchor="w", 
                                command=lambda s=sheet, i=idx: [self.load_student_into_ui(s, i), popup.destroy()])
            btn.pack(fill="x", pady=2)

    def calculate_gpa(self, grades, credits=None):
        # Standard 4.0 Scale
        GRADE_POINTS = {
            'A+': 4.00, 'A': 4.00, 'A-': 3.70,
            'B+': 3.30, 'B': 3.00, 'B-': 2.70,
            'C+': 2.30, 'C': 2.00, 'C-': 1.70,
            'D+': 1.30, 'D': 1.00, 'E': 0.00,
            'F': 0.0
        }
        
        # If credits provided, calculate weighted GPA
        if credits and len(credits) == len(grades):
            total_quality_points = 0.0
            total_credits = 0.0
            
            for i, g in enumerate(grades):
                g_clean = str(g).strip().upper()
                if g_clean in GRADE_POINTS and i < len(credits) and credits[i] is not None:
                    credit_val = credits[i] if isinstance(credits[i], (int, float)) else float(credits[i])
                    grade_points = GRADE_POINTS[g_clean]
                    total_quality_points += grade_points * credit_val
                    total_credits += credit_val
            
            if total_credits == 0: return 0.00
            return total_quality_points / total_credits
        else:
            # Simple average (unweighted) if no credits
            total_points = 0.00
            count = 0
            
            for g in grades:
                g_clean = str(g).strip().upper()
                if g_clean in GRADE_POINTS:
                    total_points += GRADE_POINTS[g_clean]
                    count += 1
            
            if count == 0: return 0.00
            return total_points / count


    def calculate_class(self, gpa):
        if gpa >= 3.70:
            return "First Class"
        elif gpa >= 3.30:
            return "Second Class (Upper Division)"
        elif gpa >= 3.00:
            return "Second Class (Lower Division)"
        elif gpa >= 2.00:
            return "Pass"
        else:
            return "Fail"


    def calculate_class(self, gpa):
        if gpa >= 3.70:
            return "First Class"
        elif gpa >= 3.30:
            return "Second Class (Upper Division)"
        elif gpa >= 3.00:
            return "Second Class (Lower Division)"
        elif gpa >= 2.00:
            return "Pass"
        else:
            return "Fail"

    def load_student_into_ui(self, sheet_name, idx):
        self.current_sheet_name = sheet_name
        self.current_student_idx = idx
        self.lbl_current_sheet.configure(text=f"Sheet: {sheet_name}")
        
        df = self.all_sheets_data[sheet_name]
        data = df.loc[idx]
        
        name_col = next((c for c in df.columns if "name" in c.lower()), None)
        name = data[name_col] if name_col else "Unknown"
        reg = data['Registration Number']
        
        self.lbl_name.configure(text=f"Name: {name}")
        self.lbl_reg.configure(text=f"Reg No: {reg}")
        
        self.btn_delete.configure(state="normal")
        self.btn_edit.configure(state="normal")
        self.btn_print.configure(state="normal")
        self.tree.delete(*self.tree.get_children())
        self.current_subject_rows = []
        
        valid_subjects = self.subject_columns_per_sheet.get(sheet_name, [])
        subject_credits = self.subject_credits_per_sheet.get(sheet_name, {})
        
        student_grades = []
        student_credits = []
        
        for i, col in enumerate(valid_subjects, start=1):
            val = data[col]
            display_val = val if pd.notna(val) and str(val).strip() != "" else "-"
            
            # Get credit value for this subject
            credit_val = subject_credits.get(col, None)
            
            # Calculate points for display
            points = "-"
            GRADE_POINTS = {
                'A+': 4.00, 'A': 4.00, 'A-': 3.70,
                'B+': 3.30, 'B': 3.00, 'B-': 2.70,
                'C+': 2.30, 'C': 2.00, 'C-': 1.70,
                'D+': 1.30, 'D': 1.00, 'E': 0.00
            }
            if str(display_val).strip().upper() in GRADE_POINTS:
                points = GRADE_POINTS[str(display_val).strip().upper()]
                student_grades.append(display_val)
                student_credits.append(credit_val)
            
            self.current_subject_rows.append({
                "s_no": i,
                "subject": col,
                "grade": display_val,
                "points": points
            })

        self.refresh_subject_tree()
        self.entry_subject_search.configure(state="normal")
        self.btn_subject_filter.configure(state="normal")
        self.subject_search_var.set("")

        # Calculate and Show GPA (with credits if available)
        gpa = self.calculate_gpa(student_grades, student_credits if any(c is not None for c in student_credits) else None)
        # Truncate to 2 decimal places (no rounding)
        truncated_gpa = math.floor(gpa * 100) / 100
        self.lbl_gpa_value.configure(text=f"{truncated_gpa:.2f}")

        # Calculate and Show Class Awarded
        class_val = self.calculate_class(gpa)
        self.lbl_class_value.configure(text=class_val)


        # Calculate and Show Class Awarded
        class_val = self.calculate_class(gpa)
        self.lbl_class_value.configure(text=class_val)


    def print_student_transcript(self):
        if self.current_student_idx is None: return
        
        name = self.lbl_name.cget("text").replace("Name: ", "")
        reg = self.lbl_reg.cget("text").replace("Reg No: ", "")
        sheet = self.lbl_current_sheet.cget("text").replace("Sheet: ", "")
        date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        df = self.all_sheets_data[self.current_sheet_name]
        row_data = df.loc[self.current_student_idx]
        
        # Get calculated GPA and Class from UI (more accurate)
        gpa_display = self.lbl_gpa_value.cget("text")
        class_display = self.lbl_class_value.cget("text")
        
        # Use UI values if available, otherwise fallback to Excel data
        if gpa_display != "-":
            gpa = gpa_display
        else:
            gpa = row_data.get('GPA', '-')
            if pd.isna(gpa): gpa = "-"
            
        if class_display != "-":
            cls = class_display
        else:
            cls = row_data.get('Class Awarded', '-')
            if pd.isna(cls): cls = "-"
        
        eff_date = row_data.get('Effective Date', '-')
        if pd.isna(eff_date): eff_date = "-"
        
        table_rows = ""
        for item in self.tree.get_children():
            vals = self.tree.item(item)['values']
            s_no = vals[0]
            subject = vals[1]
            grade = vals[2]
            table_rows += f"<tr><td style='padding:8px; border-bottom:1px solid #ddd; text-align:center; color:#666;'>{s_no}</td><td style='padding:8px; border-bottom:1px solid #ddd;'>{subject}</td><td style='padding:8px; border-bottom:1px solid #ddd; text-align:center;'><b>{grade}</b></td></tr>"

        # Build summary box content conditionally
        summary_rows = ""
        if self.include_gpa_class:
            summary_rows += f"<tr><td><b>GPA:</b></td><td>{gpa}</td></tr>"
            summary_rows += f"<tr><td><b>Class Awarded:</b></td><td>{cls}</td></tr>"
        summary_rows += f"<tr><td><b>Effective Date:</b></td><td>{eff_date}</td></tr>"

        html_content = f"""
        <html>
        <head>
            <title>Result Sheet - {reg}</title>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 40px; color: #333; }}
                h1,h2 {{ text-align: center; margin-bottom: 5px; }}
                h3 {{ text-align: center; color: #666; margin-top: 0; }}
                .header-box {{ display: flex; justify-content: space-between; border-top: 2px solid #333; border-bottom: 2px solid #333; padding: 20px 0; margin-top: 20px; }}
                .header-left, .header-right {{ width: 48%; }}
                .info-row {{ margin-bottom: 8px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 30px; }}
                th {{ background-color: #f2f2f2; text-align: left; padding: 12px; border-bottom: 2px solid #aaa; }}
                .summary-box {{ margin-top: 30px; padding: 15px; background-color: #f9f9f9; border: 1px solid #ddd; }}
                .page-footer {{ text-align: center; font-size: 12px; color: #888; padding: 10px 0; border-top: 1px solid #eee; }}
                .page-footer-dev {{ text-align: right;font-size: 7px; color: #b8b8b8; margin-top: 5px; }}
                
                @media print {{{{
                    thead {{ display: table-header-group; }}
                    tfoot {{ display: table-footer-group; }}
                    tr {{ page-break-inside: avoid; }}
                    .summary-box {{ page-break-inside: avoid; }}
                }}}}
            </style>
        </head>
        <body>
            <h2>SAB Campus of Chartered Accountants Sri Lanka </h2>
            <h1>Student Result Sheet</h1>
            <h3>{sheet}</h3>
            <div class="header-box">
                <div class="header-left">
                    <div class="info-row"><b>Name:</b> {name}</div>
                    <div class="info-row"><b>Registration No:</b> {reg}</div>
                </div>
                <div class="header-right">
                    <div class="info-row"><b>Date Issued:</b> {date_str}</div>
                </div>
            </div>
            <table>
                <thead>
                    <tr>
                        <th width="10%" style="text-align:center;">#</th>
                        <th width="70%">Subject</th>
                        <th width="20%" style="text-align:center;">Grade</th>
                    </tr>
                </thead>
                <tfoot>
                    <tr>
                        <td colspan="3" class="page-footer">
                            Generated by Student Result System
                            <div class="page-footer-dev">Dev@Salinda</div>
                        </td>
                    </tr>
                </tfoot>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <div class="summary-box">
                <table style="margin-top:0; width:50%">
                    {summary_rows}
                </table>
            </div>
            <script>window.print();</script>
        </body>
        </html>
        """
        with tempfile.NamedTemporaryFile('w', delete=False, suffix='.html', encoding='utf-8') as f:
            f.write(html_content)
            filepath = f.name
        webbrowser.open('file://' + filepath)

    def print_full_sheet(self):
        if not self.current_sheet_name: return
        df = self.all_sheets_data[self.current_sheet_name]
        valid_subjects = self.subject_columns_per_sheet.get(self.current_sheet_name, [])
        
        cols_to_show = ['Registration Number']
        name_col = next((c for c in df.columns if "name" in c.lower()), None)
        if name_col: cols_to_show.append(name_col)
        cols_to_show.extend(valid_subjects)
        
        html = df[cols_to_show].to_html(index=False, classes='clean-table', border=1)
        full_html = f"<html><body><h2>Result Sheet: {self.current_sheet_name}</h2>{html}</body></html>"
        
        with tempfile.NamedTemporaryFile('w', delete=False, suffix='.html', encoding='utf-8') as f:
            f.write(full_html)
            filepath = f.name
        webbrowser.open('file://' + filepath)

    def edit_grade(self):
        sel = self.tree.selection()
        if not sel: 
            messagebox.showwarning("Select", "Please select a subject row to edit.")
            return
        item = self.tree.item(sel)
        
        vals = item['values']
        s_no = vals[0]
        subject = vals[1]
        old_grade = vals[2]
        
        initial = old_grade if old_grade != "-" else ""
        
        # Use CTkInputDialog for better look
        dialog = ctk.CTkInputDialog(text=f"Enter grade for:\n{subject}", title="Edit Grade")
        new_grade = dialog.get_input()
        
        if new_grade is not None:
            df = self.all_sheets_data[self.current_sheet_name]
            df.at[self.current_student_idx, subject] = new_grade
            display_val = new_grade if new_grade.strip() != "" else "-"
            
            GRADE_POINTS = {
                'A+': 4.00, 'A': 4.00, 'A-': 3.70,
                'B+': 3.30, 'B': 3.00, 'B-': 2.70,
                'C+': 2.30, 'C': 2.00, 'C-': 1.70,
                'D+': 1.30, 'D': 1.00, 'E': 0.00
            }
            grade_clean = str(display_val).strip().upper()
            points = GRADE_POINTS.get(grade_clean, "-")
            
            self.tree.item(sel, values=(s_no, subject, display_val, points))

            for row in self.current_subject_rows:
                if row["subject"] == subject:
                    row["grade"] = display_val
                    row["points"] = points
                    break

            self.filter_subjects()
            
            reg_no = df.at[self.current_student_idx, 'Registration Number']
            if self.current_sheet_name not in self.pending_changes: self.pending_changes[self.current_sheet_name] = {}
            if reg_no not in self.pending_changes[self.current_sheet_name]: self.pending_changes[self.current_sheet_name][reg_no] = {}
            self.pending_changes[self.current_sheet_name][reg_no][subject] = new_grade
            messagebox.showinfo("Pending", "Grade updated in app. Click SAVE to commit.")

    def delete_student(self):
        if self.current_student_idx is None: return
        if messagebox.askyesno("Confirm", f"Delete student from {self.current_sheet_name}?"):
            df = self.all_sheets_data[self.current_sheet_name]
            reg_no = df.at[self.current_student_idx, 'Registration Number']
            self.all_sheets_data[self.current_sheet_name] = df.drop(index=self.current_student_idx).reset_index(drop=True)
            
            if self.current_sheet_name not in self.pending_deletes: self.pending_deletes[self.current_sheet_name] = []
            self.pending_deletes[self.current_sheet_name].append(reg_no)
            self.clear_ui()
            messagebox.showinfo("Pending", "Deleted from App. Click SAVE to commit.")

    def smart_save(self):
        if not self.pending_changes and not self.pending_deletes:
            messagebox.showinfo("Info", "No changes to save.")
            return
        try:
            wb = openpyxl.load_workbook(self.file_path)
            affected_sheets = set(self.pending_changes.keys()).union(set(self.pending_deletes.keys()))
            for sheet_name in affected_sheets:
                if sheet_name not in wb.sheetnames: continue
                ws = wb[sheet_name]
                
                header_row = None
                reg_col_idx = None
                for r in range(1, 15):
                    for c in range(1, ws.max_column + 1):
                        val = str(ws.cell(row=r, column=c).value).lower()
                        if "registration" in val and ("no" in val or "num" in val):
                            header_row = r
                            reg_col_idx = c
                            break
                    if header_row: break
                if not header_row: continue

                subject_col_map = {}
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(row=header_row, column=col).value
                    val_below = ws.cell(row=header_row + 1, column=col).value
                    if val_below: subject_col_map[str(val_below).strip()] = col
                    elif val: subject_col_map[str(val).strip()] = col

                reg_row_map = {}
                for row in range(header_row + 1, ws.max_row + 1):
                    val = ws.cell(row=row, column=reg_col_idx).value 
                    if val: reg_row_map[str(val).strip()] = row
                
                if sheet_name in self.pending_changes:
                    for reg_no, changes in self.pending_changes[sheet_name].items():
                        if reg_no in reg_row_map:
                            row_idx = reg_row_map[reg_no]
                            for subject, new_val in changes.items():
                                if subject in subject_col_map:
                                    ws.cell(row=row_idx, column=subject_col_map[subject]).value = new_val

                if sheet_name in self.pending_deletes:
                    rows_to_del = [reg_row_map[r] for r in self.pending_deletes[sheet_name] if r in reg_row_map]
                    for r_idx in sorted(rows_to_del, reverse=True):
                        ws.delete_rows(r_idx)

            wb.save(self.file_path)
            self.pending_changes = {}
            self.pending_deletes = {}
            messagebox.showinfo("Success", "Changes saved to Excel file!")
        except PermissionError:
            messagebox.showerror("Error", "Please close the Excel file and try again!")
        except Exception as e:
            messagebox.showerror("Error", f"Save failed: {e}")

    def clear_ui(self):
        self.lbl_name.configure(text="Name: -")
        self.lbl_reg.configure(text="Reg No: -")
        self.lbl_current_sheet.configure(text="Sheet: -")
        self.tree.delete(*self.tree.get_children())
        self.current_subject_rows = []
        self.current_student_idx = None
        self.btn_edit.configure(state="disabled")
        self.btn_delete.configure(state="disabled")
        self.btn_print.configure(state="disabled")
        self.entry_subject_search.configure(state="disabled")
        self.btn_subject_filter.configure(state="disabled")
        self.subject_search_var.set("")

    def refresh_subject_tree(self, filter_term=None):
        self.tree.delete(*self.tree.get_children())
        if not self.current_subject_rows:
            return
        rows = self.current_subject_rows
        if filter_term:
            ft = filter_term.lower()
            rows = [row for row in rows if ft in str(row["subject"]).lower()]
        for row in rows:
            self.tree.insert("", "end", values=(row["s_no"], row["subject"], row["grade"], row["points"]))

    def filter_subjects(self):
        if not self.current_subject_rows:
            return
        term = self.subject_search_var.get().strip()
        self.refresh_subject_tree(term if term else None)

    def open_subject_results_window(self):
        if not self.all_sheets_data:
            messagebox.showinfo("Info", "Please load a workbook first.")
            return

        if self.subject_results_window and self.subject_results_window.winfo_exists():
            self.subject_results_window.focus()
            return

        window = ctk.CTkToplevel(self)
        window.title("Batch Subject Results")
        window.geometry("1050x650")
        window.transient(self)
        window.grab_set()
        self.subject_results_window = window

        def on_close():
            self.subject_results_window = None
            window.destroy()

        window.protocol("WM_DELETE_WINDOW", on_close)

        batches = list(self.all_sheets_data.keys())
        initial_batch = self.current_sheet_name if self.current_sheet_name in batches else (batches[0] if batches else "")
        batch_var = ctk.StringVar(value=initial_batch)
        subject_var = ctk.StringVar(value="")
        subject_menu = None
        results_tree = None
        btn_show = None
        btn_print_results = None
        info_label = None
        current_results = []

        GRADE_POINTS = {
            'A+': 4.00, 'A': 4.00, 'A-': 3.70,
            'B+': 3.30, 'B': 3.00, 'B-': 2.70,
            'C+': 2.30, 'C': 2.00, 'C-': 1.70,
            'D+': 1.30, 'D': 1.00, 'E': 0.00
        }

        def update_info(text, color="gray"):
            if info_label:
                info_label.configure(text=text, text_color=color)

        def clear_results_tree():
            if results_tree:
                for item in results_tree.get_children():
                    results_tree.delete(item)

        def update_show_button_state():
            if btn_show:
                btn_show.configure(state="normal" if subject_var.get() else "disabled")

        def update_subject_menu(selected_batch):
            subjects = self.subject_columns_per_sheet.get(selected_batch, [])
            if subject_menu:
                if subjects:
                    subject_menu.configure(values=subjects, state="normal")
                    if subject_var.get() not in subjects:
                        subject_var.set(subjects[0])
                else:
                    subject_menu.configure(values=["No subjects available"], state="disabled")
                    subject_var.set("")
            update_show_button_state()

        def handle_batch_change(selected_batch):
            batch_var.set(selected_batch)
            update_subject_menu(selected_batch)
            clear_results_tree()
            update_info(f"Select a subject to view results for {selected_batch}.", "#FFC107")

        def handle_subject_change(selected_subject):
            subject_var.set(selected_subject)
            update_show_button_state()

        def populate_tree():
            clear_results_tree()
            batch = batch_var.get()
            subject = subject_var.get()
            current_results.clear()
            if not batch or not subject:
                update_info("Select a batch and subject to view results.", "#FFC107")
                if btn_print_results:
                    btn_print_results.configure(state="disabled")
                return

            df = self.all_sheets_data.get(batch)
            if df is None or subject not in df.columns:
                update_info("Selected subject not found in this batch.", "#EF5350")
                return

            name_col = next((c for c in df.columns if "name" in c.lower()), None)
            rows_added = 0

            for idx, row in df.iterrows():
                reg_val = row.get('Registration Number', "-")
                reg_display = "-" if pd.isna(reg_val) else str(reg_val).strip()

                if name_col:
                    name_val = row.get(name_col, "-")
                    name_display = "-" if pd.isna(name_val) else str(name_val).strip()
                else:
                    name_display = "-"

                grade_val = row.get(subject, "-")
                grade_display = "-"
                if pd.notna(grade_val) and str(grade_val).strip():
                    grade_display = str(grade_val).strip()

                grade_clean = grade_display.upper()
                points = GRADE_POINTS.get(grade_clean, "-")

                row_values = (idx + 1, reg_display, name_display, grade_display, points)
                if results_tree:
                    results_tree.insert("", "end", values=row_values)
                current_results.append({
                    "s_no": idx + 1,
                    "reg_no": reg_display,
                    "name": name_display,
                    "grade": grade_display,
                    "points": points
                })
                rows_added += 1

            if rows_added:
                update_info(f"{rows_added} result(s) shown for {subject} in {batch}.", "#A5D6A7")
                if btn_print_results:
                    btn_print_results.configure(state="normal")
            else:
                update_info("No candidates found for this selection.", "#FFAB91")
                if btn_print_results:
                    btn_print_results.configure(state="disabled")

        def print_results():
            if not current_results:
                messagebox.showinfo("Info", "Load results before printing.")
                return
            batch = batch_var.get()
            subject = subject_var.get()
            date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            table_rows = ""
            for row in current_results:
                table_rows += (
                    f"<tr>"
                    f"<td style='padding:8px;border-bottom:1px solid #ddd;text-align:center;color:#666;'>{row['s_no']}</td>"
                    f"<td style='padding:8px;border-bottom:1px solid #ddd;text-align:center;'>{row['reg_no']}</td>"
                    f"<td style='padding:8px;border-bottom:1px solid #ddd;'>{row['name']}</td>"
                    f"<td style='padding:8px;border-bottom:1px solid #ddd;text-align:center;'><b>{row['grade']}</b></td>"
                    f"<td style='padding:8px;border-bottom:1px solid #ddd;text-align:center;'>{row['points']}</td>"
                    f"</tr>"
                )
            html_content = f"""
            <html>
            <head>
                <title>{batch} - {subject}</title>
                <style>
                    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 40px; color: #333; }}
                    h1,h2 {{ text-align: center; margin-bottom: 5px; }}
                    .meta {{ text-align: center; margin-bottom: 25px; color:#666; }}
                    table {{ width: 100%; border-collapse: collapse; }}
                    th {{ background-color: #f2f2f2; padding: 12px; border-bottom: 2px solid #aaa; }}
                    td {{ font-size: 14px; }}
                    .page-footer {{ text-align: center; font-size: 12px; color: #888; padding: 10px 0; border-top: 1px solid #eee; margin-top:30px; }}
                    @media print {{
                        thead {{ display: table-header-group; }}
                        tfoot {{ display: table-footer-group; }}
                        tr {{ page-break-inside: avoid; }}
                    }}
                </style>
            </head>
            <body>
                <h1>Subject Results</h1>
                <div class="meta">
                    <div><b>Batch:</b> {batch}</div>
                    <div><b>Subject:</b> {subject}</div>
                    <div><b>Generated:</b> {date_str}</div>
                </div>
                <table>
                    <thead>
                        <tr>
                            <th width="8%">#</th>
                            <th width="20%">Registration No</th>
                            <th width="42%">Name</th>
                            <th width="15%">Grade</th>
                            <th width="15%">Points</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows}
                    </tbody>
                </table>
                <div class="page-footer">Generated by Student Result System</div>
                <script>window.print();</script>
            </body>
            </html>
            """
            with tempfile.NamedTemporaryFile('w', delete=False, suffix='.html', encoding='utf-8') as f:
                f.write(html_content)
                filepath = f.name
            webbrowser.open('file://' + filepath)

        control_frame = ctk.CTkFrame(window)
        control_frame.pack(fill="x", padx=20, pady=20)

        ctk.CTkLabel(control_frame, text="Batch:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=(0, 10))

        batch_menu = ctk.CTkOptionMenu(
            control_frame,
            values=batches if batches else ["-"],
            command=handle_batch_change,
            variable=batch_var,
            width=220
        )
        batch_menu.pack(side="left", padx=(0, 20))

        ctk.CTkLabel(control_frame, text="Subject:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=(0, 10))

        initial_subjects = self.subject_columns_per_sheet.get(batch_var.get(), []) if batch_var.get() else []
        initial_subject_value = initial_subjects[0] if initial_subjects else ""
        subject_var.set(initial_subject_value)

        subject_menu = ctk.CTkOptionMenu(
            control_frame,
            values=initial_subjects if initial_subjects else ["No subjects available"],
            command=handle_subject_change,
            variable=subject_var,
            width=320,
            state="normal" if initial_subjects else "disabled"
        )
        subject_menu.pack(side="left", padx=(0, 20))

        btn_show = ctk.CTkButton(control_frame, text="Show Results", command=populate_tree, state="normal" if initial_subjects else "disabled")
        btn_show.pack(side="left", padx=(0, 10))

        btn_print_results = ctk.CTkButton(control_frame, text="Print Results", command=print_results, state="disabled")
        btn_print_results.pack(side="left")

        results_frame = ctk.CTkFrame(window)
        results_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        tree_scroll = ctk.CTkScrollbar(results_frame)
        tree_scroll.pack(side="right", fill="y")

        columns = ("SNo", "RegNo", "Name", "Grade", "Points")
        results_tree = ttk.Treeview(results_frame, columns=columns, show="headings", yscrollcommand=tree_scroll.set)
        results_tree.heading("SNo", text="#")
        results_tree.heading("RegNo", text="Registration No")
        results_tree.heading("Name", text="Name")
        results_tree.heading("Grade", text="Grade")
        results_tree.heading("Points", text="Points")

        results_tree.column("SNo", width=70, anchor="center")
        results_tree.column("RegNo", width=180, anchor="center")
        results_tree.column("Name", width=260)
        results_tree.column("Grade", width=120, anchor="center")
        results_tree.column("Points", width=120, anchor="center")

        results_tree.pack(fill="both", expand=True, padx=2, pady=2)
        tree_scroll.configure(command=results_tree.yview)

        info_label = ctk.CTkLabel(window, text="", text_color="gray")
        info_label.pack(fill="x", padx=20, pady=(0, 15))

        update_subject_menu(batch_var.get())
        populate_tree()

if __name__ == "__main__":
    app = StudentResultApp()
    app.mainloop()