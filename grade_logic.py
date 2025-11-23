import pandas as pd
import openpyxl
import re
import math

# Constants
GRADE_POINTS = {
    'A+': 4.00, 'A': 4.00, 'A-': 3.70,
    'B+': 3.30, 'B': 3.00, 'B-': 2.70,
    'C+': 2.30, 'C': 2.00, 'C-': 1.70,
    'D+': 1.30, 'D': 1.00, 'E': 0.00,
    'F': 0.0
}

EXCLUDE_KEYWORDS = [
    'serial', 'registration', 'name', 'general', 'special', 'drop', 'batch', 'transfer',
    'finance clearance', 'results confirmation', 'gpa', 'class', 'effective', 'total no', 
    'nq ese', 'ab ese', 'repeat ese', 'hold', 'results released', 'range of marks', 'grade points', 'letter grade',
    'pendings', 'pending'
]

GRADE_COLS = ['A+', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'E', 'AB', 'B and Above', 'C and above', 'A+/A/A-', 'EX', '-', 'C-/D+']

def calculate_gpa(grades, credits=None):
    """Calculates GPA based on grades and optional credits."""
    if credits and len(credits) == len(grades):
        total_quality_points = 0.0
        total_credits = 0.0
        
        for i, g in enumerate(grades):
            g_clean = str(g).strip().upper()
            if g_clean in GRADE_POINTS and i < len(credits) and credits[i] is not None:
                credit_val = credits[i] if isinstance(credits[i], (int, float)) else float(credits[i])
                grade_points = GRADE_POINTS[g_clean]
                total_quality_points += grade_points * credit_val
                total_credits += credit_val
        
        if total_credits == 0: return 0.00
        return total_quality_points / total_credits
    else:
        # Simple average (unweighted) if no credits
        total_points = 0.00
        count = 0
        
        for g in grades:
            g_clean = str(g).strip().upper()
            if g_clean in GRADE_POINTS:
                total_points += GRADE_POINTS[g_clean]
                count += 1
        
        if count == 0: return 0.00
        return total_points / count

def calculate_class(gpa):
    """Determines class based on GPA."""
    if gpa >= 3.70:
        return "First Class"
    elif gpa >= 3.30:
        return "Second Class (Upper Division)"
    elif gpa >= 3.00:
        return "Second Class (Lower Division)"
    elif gpa >= 2.00:
        return "Pass"
    else:
        return "Fail"

def load_workbook_data(file_path):
    """Loads workbook and processes sheets."""
    try:
        all_sheets_raw = pd.read_excel(file_path, sheet_name=None, header=None)
        all_sheets_data = {}
        subject_columns_per_sheet = {}
        subject_credits_per_sheet = {}
        
        for sheet_name, raw_df in all_sheets_raw.items():
            if len(raw_df) < 1: continue
            
            # Find Header Row
            header_row_idx = -1
            for i in range(min(10, len(raw_df))):
                row_vals = raw_df.iloc[i].astype(str).str.lower().tolist()
                if any("registration" in x for x in row_vals) and (any("no" in x for x in row_vals) or any("number" in x for x in row_vals)):
                    header_row_idx = i
                    break
            
            if header_row_idx == -1: continue
                
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx)
            
            # Merge Subject Row (Row below header)
            if len(df) > 0:
                subject_row = df.iloc[0]
                new_columns = list(df.columns)
                for i in range(len(new_columns)):
                    if i >= 4: 
                        val = subject_row.iloc[i]
                        if pd.notna(val) and str(val).strip() != "":
                            new_columns[i] = str(val).strip()
                df.columns = new_columns
                df = df.drop(0).reset_index(drop=True)

            # Normalize Columns
            df.columns = [str(c).strip() for c in df.columns]
            reg_col_found = False
            for col in df.columns:
                if "registration" in col.lower() or "reg no" in col.lower():
                    df.rename(columns={col: 'Registration Number'}, inplace=True)
                    reg_col_found = True
                    break
            
            if not reg_col_found: continue
            
            # Load ONLY subject columns matching the pattern (BSAA XXXXX Subject Name)
            subject_cols = []
            
            for idx, col in enumerate(df.columns):
                col_str = str(col).strip()
                
                # Only include columns that match the subject code pattern
                # Pattern: 2-6 uppercase letters + space + 5 digits + space + subject name
                # Example: BSAA 11013 Financial Accounting
                is_subject = bool(re.match(r'^[A-Z]{2,6}\s+\d{5}\s+', col_str))
                
                if is_subject:
                    subject_cols.append(col_str)
            
            subject_columns_per_sheet[sheet_name] = subject_cols
            
            # Extract credit values
            subject_credits = {}
            for col_name in subject_cols:
                subject_code = str(col_name).strip()
                digits = re.findall(r'\d', subject_code)
                if digits:
                    credit_val = int(digits[-1])
                    if credit_val > 0:
                        subject_credits[col_name] = float(credit_val)
            
            subject_credits_per_sheet[sheet_name] = subject_credits
            all_sheets_data[sheet_name] = df

        return all_sheets_data, subject_columns_per_sheet, subject_credits_per_sheet
        
    except Exception as e:
        return None, None, None

def save_changes_to_excel(file_path, pending_changes, pending_deletes):
    """Saves pending changes to the Excel file."""
    try:
        wb = openpyxl.load_workbook(file_path)
        affected_sheets = set(pending_changes.keys()).union(set(pending_deletes.keys()))
        
        for sheet_name in affected_sheets:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            
            header_row = None
            reg_col_idx = None
            for r in range(1, 15):
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=r, column=c).value).lower()
                    if "registration" in val and ("no" in val or "num" in val):
                        header_row = r
                        reg_col_idx = c
                        break
                if header_row: break
            if not header_row: continue

            subject_col_map = {}
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col).value
                val_below = ws.cell(row=header_row + 1, column=col).value
                if val_below: subject_col_map[str(val_below).strip()] = col
                elif val: subject_col_map[str(val).strip()] = col

            reg_row_map = {}
            for row in range(header_row + 1, ws.max_row + 1):
                val = ws.cell(row=row, column=reg_col_idx).value 
                if val: reg_row_map[str(val).strip()] = row
            
            if sheet_name in pending_changes:
                for reg_no, changes in pending_changes[sheet_name].items():
                    if reg_no in reg_row_map:
                        row_idx = reg_row_map[reg_no]
                        for subject, new_val in changes.items():
                            if subject in subject_col_map:
                                ws.cell(row=row_idx, column=subject_col_map[subject]).value = new_val

            if sheet_name in pending_deletes:
                rows_to_del = [reg_row_map[r] for r in pending_deletes[sheet_name] if r in reg_row_map]
                for r_idx in sorted(rows_to_del, reverse=True):
                    ws.delete_rows(r_idx)

        wb.save(file_path)
        return True, "Changes saved successfully!"
    except PermissionError:
        return False, "Please close the Excel file and try again!"
    except Exception as e:
        return False, f"Save failed: {str(e)}"
